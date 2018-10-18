from categories import Categories
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用


categories=Categories().getcategories()
paths=[]
global_wb=workbook.Workbook()
global_ws=global_wb.active

last_path="D:\\Works\\datas\\outputFiles\\"+"百度知道体检知识.xlsx"
for cate in categories:
    path="D:\\Works\\datas\\outputFiles\\"+cate+"百度知道体检知识.xlsx"
    paths.append(path)
    wb = load_workbook(path)
    a_sheet = wb.get_sheet_by_name('Sheet')
    rows = a_sheet.max_row
    columns = a_sheet.max_column
    now_row_number =0
    for row in a_sheet.rows:
        row_data=[]
        now_row_number+=1
        if(now_row_number==1):
            continue
        for cell in row:
            row_data.append(cell.value)
        global_ws.append(row_data)

global_wb.save(last_path)